"""
Exporter Module
Generates enriched output files (CSV/XLSX with multi-tab support & cell highlighting), evidence JSON, and review CSV list.
Highlights newly enriched attribute cells in soft green (#D1FAE5) and low-confidence flagged cells in soft amber (#FEF3C7).
"""

import json
import logging
import os
from typing import Dict, Any, List
import pandas as pd

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_enrichment_results(
    enriched_df: pd.DataFrame,
    evidence_records: List[Dict[str, Any]],
    output_path: str,
    evidence_json_path: str,
    review_csv_path: str,
    confidence_threshold: int = 75
) -> Dict[str, Any]:
    """
    Exports all required output formats with visual cell highlighting for enriched cells:
      1. Main enriched file (CSV or XLSX with Enriched + Evidence tabs and green highlighting for enriched cells).
      2. Evidence JSON artifact.
      3. Review CSV artifact for low-confidence fills (< threshold).
    """
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    os.makedirs(os.path.dirname(os.path.abspath(evidence_json_path)), exist_ok=True)
    os.makedirs(os.path.dirname(os.path.abspath(review_csv_path)), exist_ok=True)

    # 1. Prepare Evidence DataFrame
    evidence_df = pd.DataFrame(evidence_records)
    if evidence_df.empty:
        evidence_df = pd.DataFrame(columns=[
            "Row_Index", "SKU_ID", "Product_Name", "Attribute", 
            "Filled_Value", "Unit", "Source", "Source_Type", 
            "URL", "Confidence", "Flagged_For_Review", "Evidence_Note"
        ])

    # Filter review list (flagged items or confidence < threshold)
    review_df = pd.DataFrame()
    if not evidence_df.empty and "Confidence" in evidence_df.columns:
        review_df = evidence_df[
            (evidence_df["Flagged_For_Review"] == True) | 
            (evidence_df["Confidence"] < confidence_threshold)
        ].copy()

    # 2. Export Main File
    is_excel = output_path.endswith(".xlsx") or output_path.endswith(".xls")
    
    if is_excel and OPENPYXL_AVAILABLE:
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Enriched Sheet
        ws_enriched = wb.create_sheet(title="Enriched")
        
        # Write Headers
        headers = list(enriched_df.columns)
        ws_enriched.append(headers)

        # Header styling
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_enriched.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Build column index lookup
        col_name_to_idx = {col: i + 1 for i, col in enumerate(headers)}

        # Write Data Rows
        for row_idx, row in enriched_df.iterrows():
            excel_row = row_idx + 2
            row_vals = [row[c] for c in headers]
            ws_enriched.append(row_vals)

        # Color Formatting Fills for Enriched Cells
        green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Soft Emerald Green
        green_font = Font(name="Calibri", size=11, bold=True, color="065F46")

        amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Soft Amber
        amber_font = Font(name="Calibri", size=11, bold=True, color="92400E")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for rec in evidence_records:
            r_idx = rec.get("Row_Index")  # 1-indexed
            attr_col = rec.get("Attribute")
            flagged = rec.get("Flagged_For_Review", False)

            if r_idx and attr_col in col_name_to_idx:
                excel_row = r_idx + 1  # Header is row 1, Row_Index 1 is row 2
                c_idx = col_name_to_idx[attr_col]

                cell = ws_enriched.cell(row=excel_row, column=c_idx)
                if flagged:
                    cell.fill = amber_fill
                    cell.font = amber_font
                else:
                    cell.fill = green_fill
                    cell.font = green_font
                cell.border = thin_border

        # Evidence Sheet
        ws_evidence = wb.create_sheet(title="Evidence")
        ws_evidence.append(list(evidence_df.columns))
        for _, row in evidence_df.iterrows():
            ws_evidence.append(list(row))

        # Header style for Evidence
        for col_idx in range(1, len(evidence_df.columns) + 1):
            cell = ws_evidence.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        # Auto-fit Column Widths
        for sheet in [ws_enriched, ws_evidence]:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_path)
        logger.info(f"Saved Excel output with green cell highlighting for enriched values: {output_path}")

    else:
        if is_excel:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                enriched_df.to_excel(writer, sheet_name="Enriched", index=False)
                evidence_df.to_excel(writer, sheet_name="Evidence", index=False)
        else:
            enriched_df.to_csv(output_path, index=False)
        logger.info(f"Saved enriched output: {output_path}")

    # 3. Export Evidence JSON
    with open(evidence_json_path, "w", encoding="utf-8") as f:
        json.dump(evidence_records, f, indent=2, ensure_ascii=False)
    logger.info(f"Saved Evidence JSON artifact: {evidence_json_path}")

    # 4. Export Review CSV
    if not review_df.empty:
        review_df.to_csv(review_csv_path, index=False)
    else:
        pd.DataFrame(columns=[
            "Row_Index", "SKU_ID", "Product_Name", "Attribute", 
            "Filled_Value", "Unit", "Source", "Source_Type", 
            "URL", "Confidence", "Flagged_For_Review", "Evidence_Note"
        ]).to_csv(review_csv_path, index=False)
    logger.info(f"Saved Review CSV artifact ({len(review_df)} items): {review_csv_path}")

    return {
        "output_path": output_path,
        "evidence_json_path": evidence_json_path,
        "review_csv_path": review_csv_path,
        "total_enriched_rows": len(enriched_df),
        "total_evidence_entries": len(evidence_records),
        "flagged_for_review": len(review_df)
    }
